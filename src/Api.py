from flask import Flask, request, jsonify,send_file
import pandas as pd
from flask_cors import CORS
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side


app = Flask(__name__)
CORS(app)

@app.route('/getData', methods=['GET'])
def upload_file():
    try:
        df = pd.read_excel("C:\\vue2\\mytest\\src\\Data\\data.xlsx")
        df = df.where(pd.notnull(df), None)
        data = df.to_dict(orient='records')
        print(data)
        return jsonify(data), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    

@app.route('/createPart', methods=['POST'])
def createPart():
    checkcolor = False ; 
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    try:
        data = request.json  
        part_name = data.get('name')  
        file_path = "C:\\vue2\\mytest\\src\\Data\\data.xlsx"

        df = pd.read_excel(file_path)
        part_rows = df[df.iloc[:, 0] == 'Part No'] 
        if not part_rows.empty:
            part_index = part_rows.index[0]  

            last_column_index = len(df.columns) - 1
            df.loc[part_index, last_column_index] = part_name 
            
            new_row = pd.DataFrame([[part_name] + [None] * (len(df.columns) - 1)], columns=df.columns)
            df = pd.concat([df, new_row], ignore_index=True)
            df.columns = [None] * len(df.columns)
            df.to_excel(file_path, index=False)

            workbook = load_workbook(file_path)
            sheet = workbook.active
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            green_fill = PatternFill(start_color='2ecc71', end_color='2ecc71', fill_type='solid')
            gray_fill = PatternFill(start_color='f0f0f0', end_color='f0f0f0', fill_type='solid')
            white_fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
            for row in range(1, sheet.max_row + 1):  
                cell = sheet.cell(row=row, column=1)  
                if(checkcolor):
                    sheet.cell(row=row, column=1).fill = yellow_fill
                    for col in range(2, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).border = thin_border
                        if sheet.cell(row=row, column=col).value is None:
                            sheet.cell(row=row, column=col).fill = gray_fill
                        else :
                            sheet.cell(row=row, column=col).fill = white_fill
                        
                if cell.value == 'Part No':
                    checkcolor=True
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).fill = yellow_fill
                        sheet.cell(row=row, column=col).border = thin_border
                    sheet.cell(row=row, column=1).fill = green_fill
            workbook.save(file_path)

        return jsonify({"message": "Part created successfully"}), 200

    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500

@app.route('/download', methods=['GET'])
def download():
    try:
        file_path = "C:\\vue2\\mytest\\src\\Data\\data.xlsx"
        return send_file(file_path, as_attachment=True)  

    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500
    
    
@app.route('/import', methods=['POST'])
def importdata():
    checkcolor = False ; 
    thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file part"}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No selected file"}), 400
        
        df = pd.read_excel(file) 
        file_path = "C:\\vue2\\mytest\\src\\Data\\data.xlsx"
        df.columns = [None] * len(df.columns)
        df.to_excel(file_path, index=False )
        workbook = load_workbook(file_path)
        sheet = workbook.active
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        green_fill = PatternFill(start_color='2ecc71', end_color='2ecc71', fill_type='solid')
        gray_fill = PatternFill(start_color='f0f0f0', end_color='f0f0f0', fill_type='solid')
        white_fill = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
        for row in range(1, sheet.max_row + 1):  
            print(row)
            cell = sheet.cell(row=row, column=1)  
            if(checkcolor):
                sheet.cell(row=row, column=1).fill = yellow_fill
                for col in range(2, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).border = thin_border
                    if sheet.cell(row=row, column=col).value is None:
                        sheet.cell(row=row, column=col).fill = gray_fill
                    else :
                        sheet.cell(row=row, column=col).fill = white_fill
                        
            if cell.value == 'Part No':
                checkcolor=True
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).fill = yellow_fill
                    sheet.cell(row=row, column=col).border = thin_border
                sheet.cell(row=row, column=1).fill = green_fill
        workbook.save(file_path)
    
        return jsonify({"message": "File imported successfully"}), 200

    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True)

